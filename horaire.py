#!/Library/Frameworks/Python.framework/Versions/3.10/bin/python3
"""
Création d'un fichier iCalendar pour l'horaire.
"""
import getopt
import os
import datetime
import sys
from zipfile import BadZipFile
import openpyxl  # type: ignore
import icalendar

def affiche_aide():
    """
        Affiche l'aide pour la commande.
    """
    print("""
    horaire.py -i <fichier_modele> -o <fichier_sortie> -h

    -i : Le chiffrier Excel contenant le modèle d'horaire.
    -o : Le fichier dans lequel sera créé l'horaire (.ics).
    -h : L'aide de la commande.
    """)

def creer_horaire(fichier_modele, fichier_sortie):
    """
    Lecture d'un modèle d'horaire (Excel) et
    création de toutes les entrées d'un cours pour la session en format iCalendar.

        Paramètres
        ----------
        fichier_modele : str
            Nom et chemin du chiffrier Excel contenant l'horaire à créer.
        fichier_sortie : str
            Nom et chemin du fichier iCalendar à créer.
    """
    cal = icalendar.Calendar()
    cal.add('prodid', '-//Mon Horaire//cegepvicto.ca//')
    cal.add('version', '2.0')

    modele = openpyxl.load_workbook(fichier_modele, data_only=True)

    f_calendrier = modele["Calendrier"]
    f_cours = modele["Cours"]

    midi = datetime.time(12, 0, 0)

    for ligne_calendrier in range(2, f_calendrier.max_row + 1):
        for l_cours in range(2, f_cours.max_row + 1):
            mode_jour = f_calendrier.cell(row=ligne_calendrier, column=3).value
            heure_debut_cours = f_cours.cell(row=l_cours, column=3).value
            mode_horaire = "AM" if heure_debut_cours < midi else "PM"

            if (
                mode_jour in ("COMPLET", mode_horaire) and
                f_calendrier.cell(row=ligne_calendrier, column=2).value ==
                f_cours.cell(row=l_cours, column=2).value
            ):
                event = icalendar.Event()
                event.add('summary', f_cours.cell(row=l_cours, column=1).value)

                date_cours = f_calendrier.cell(row=ligne_calendrier, column=1).value
                heure_debut = f_cours.cell(row=l_cours, column=3).value
                heure_fin = f_cours.cell(row=l_cours, column=4).value

                dtstart = datetime.datetime.combine(date_cours, heure_debut)
                dtend = datetime.datetime.combine(date_cours, heure_fin)

                event.add('dtstart', dtstart)
                event.add('dtend', dtend)

                emplacement = f_cours.cell(row=l_cours, column=5).value
                if emplacement:
                    event.add('location', emplacement)

                cal.add_component(event)

    # Enregistrer le fichier iCalendar
    with open(fichier_sortie, 'wb') as f:
        f.write(cal.to_ical())

def valider_parametres(fichier_modele):
    """
        Valide l'ensemble des paramètres reçus en ligne de commande.
        Vérifie que le chiffrier contient bien les critères nécessaires.

        Paramètres
        ----------
        fichier_modele : str
            Nom et chemin du chiffrier Excel contenant l'horaire à créer.

        Retour
        ------
        True si tout est valide.
    """
    parametres_valides = True

    # Validation des paramètres
    if not os.path.isfile(fichier_modele):
        print(f"Le fichier d'entrée {fichier_modele} n'existe pas.")
        parametres_valides = False

    # Vérifier si le fichier d'entrée est un chiffrier Excel
    try:
        chiffrier = openpyxl.load_workbook(fichier_modele, data_only=True)

        # Vérifier si les feuilles existent
        if "Calendrier" not in chiffrier.sheetnames:
            print("La feuille Calendrier n'existe pas.")
            parametres_valides = False
        if "Cours" not in chiffrier.sheetnames:
            print("La feuille Cours n'existe pas.")
            parametres_valides = False
    except BadZipFile:
        print(f"Le fichier d'entrée {fichier_modele} n'est pas un chiffrier Excel valide.")
        parametres_valides = False

    return parametres_valides

def main(argv):
    """
        Procédure principale
    """
    fichier_modele = ''
    fichier_sortie = ''

    currentdir = os.getcwd()

    try:
        opts, _ = getopt.getopt(argv, "hi:o:")
    except getopt.GetoptError:
        affiche_aide()
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            affiche_aide()
            sys.exit()
        elif opt == '-i':
            fichier_modele = os.path.join(currentdir, arg)
        elif opt == '-o':
            fichier_sortie = os.path.join(currentdir, arg)

    if valider_parametres(fichier_modele):
        print(f'Fichier d\'entrée est : "{fichier_modele}"')
        print(f'Fichier de sortie est : "{fichier_sortie}"')
        creer_horaire(fichier_modele, fichier_sortie)

if __name__ == "__main__":
    main(sys.argv[1:])